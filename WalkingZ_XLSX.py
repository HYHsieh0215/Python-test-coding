import os
import openpyxl
import tkinter as tk
from tkinter import filedialog
root = tk.Tk()
root.withdraw()
file_path = filedialog.askdirectory()
wb = openpyxl.Workbook()                # 建立空白ＥＸＣＥＬ活頁簿
s1 = wb.create_sheet('WalkingZ',0)        #建立工作表
print('File location is ' + file_path)
DeviceName = input("Device Name is ")
Pins_Cnt =input("Digital Pins count is ")
Path = file_path + '/' + DeviceName + '_WalkingZ_'+ Pins_Cnt + '.xlsx'
print('File name is ' + Path)
print('import tset tset_WalkZ;')
s1.cell(1,1).value='import tset tset_WalkZ;' # A1寫資料
print('vm_vector    ( $tset , AllDigPins)')
s1.cell(2,1).value='vm_vector    ( $tset , AllDigPins)' #A2寫資料
print('{')
s1.cell(3,1).value='{'  #A3寫資料
PinsStr_0 =''
countA =0
for number in range(0,int(Pins_Cnt)):
    PinsStr_0 = PinsStr_0 + "0"
print('                    > tset_WalkZ     '+ PinsStr_0 + ';//Vector ='+str(countA))
s1.cell(4,1).value='                    > tset_WalkZ     '+ PinsStr_0 + ';//Vector ='+str(countA)
countA=countA+1
for number in range(0,int(Pins_Cnt)):
    PinsStrlist = list(PinsStr_0)
    PinsStrlist_M =list(PinsStr_0)
    PinsStrlist[number] = "X"
    PinsStrlist_M[number] = "M"
    PinsStr_0_new =''
    PinsStr_M_new =''
    for number1 in range(0,int(Pins_Cnt)):
        PinsStr_0_new = PinsStr_0_new + str(PinsStrlist[number1])
        PinsStr_M_new = PinsStr_M_new + str(PinsStrlist_M[number1])
    print('repeat 20           > tset_WalkZ     ' + PinsStr_0_new + ';//Vector ='+str(countA))
    s1.cell(2*number+5,1).value='repeat 20           > tset_WalkZ     ' + PinsStr_0_new + ';//Vector ='+str(countA)
    countA=countA+1
    print('                    > tset_WalkZ     ' + PinsStr_M_new + ';//Vector ='+str(countA))
    s1.cell(2*number+6,1).value='                    > tset_WalkZ     ' + PinsStr_M_new + ';//Vector ='+str(countA)
    countA=countA+1
print('halt                > tset_WalkZ     ' + PinsStr_0 + ';//Vector ='+str(countA))
s1.cell(countA+4,1).value='halt                > tset_WalkZ     ' + PinsStr_0 + ';//Vector ='+str(countA)
countA=countA+1
if int(countA) < 64 :
    for number in range(int(countA) ,64):
        print('                    > tset_WalkZ     ' + PinsStr_0 + ';//Dummy Vector ='+str(countA))
        s1.cell(countA+4,1).value='                    > tset_WalkZ     ' + PinsStr_0 + ';//Dummy Vector ='+str(countA)
        countA=countA+1 
print('}')
s1.cell(countA+4,1).value='}'
wb.save(str(Path))
